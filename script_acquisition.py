from openpyxl import load_workbook
from openpyxl import Workbook
import os
import array
import config
import gen_EH05.gen_BB_EH05,gen_EH05.gen_BR_EH05,gen_EH05.gen_HVAC_EH05,gen_EH05.gen_PEMS_EH05
import gen_PH2.gen_AUX_PH2,gen_PH2.gen_INV_PH2,gen_PH2.gen_PEMS_PH2,gen_PH2.gen_SKID_PH2
import gen_HH.gen_AUX_HH,gen_HH.gen_BB_HH, gen_HH.gen_BR_HH,gen_HH.gen_HVAC_HH,gen_HH.gen_INV_HH,gen_HH.gen_PEMS_HH,gen_HH.gen_SKID_HH


wb = load_workbook('excel_sviluppo.xlsx')    
ws = wb.active

config.n_PI = ws['L2'].value
config.num_HH_GUI = ws['L2'].value

#array acquisition HVAC
config.n_HVAC.append(ws['B3'].value)
config.n_HVAC.append(ws['B13'].value)
config.n_HVAC.append(ws['B23'].value)
config.n_HVAC.append(ws['B33'].value)

#battery bank acquisition
config.n_battery_bank.append(ws['B4'].value)
config.n_battery_bank.append(ws['B14'].value)
config.n_battery_bank.append(ws['B24'].value)
config.n_battery_bank.append(ws['B34'].value)

#battery rack acquisition
config.n_battery_rack.append(ws['B5'].value)
config.n_battery_rack.append(ws['B15'].value)
config.n_battery_rack.append(ws['B25'].value)
config.n_battery_rack.append(ws['B35'].value)

#n CBESS acquisition
config.n_CBESS.append(ws['B6'].value)
config.n_CBESS.append(ws['B16'].value)
config.n_CBESS.append(ws['B26'].value)
config.n_CBESS.append(ws['B36'].value)

#n other inverter acquisition
config.n_other_inverter.append(ws['B7'].value)
config.n_other_inverter.append(ws['B17'].value)
config.n_other_inverter.append(ws['B27'].value)
config.n_other_inverter.append(ws['B37'].value)

#array type acquisition
config.array_type.append(ws['B2'].value)
config.array_type.append(ws['B12'].value)
config.array_type.append(ws['B22'].value)
config.array_type.append(ws['B32'].value)

#n PH acquisition
config.array_PH.append(ws['B10'].value)
config.array_PH.append(ws['B20'].value)
config.array_PH.append(ws['B30'].value)
config.array_PH.append(ws['B40'].value)

#n EH acquisition
config.array_EH.append(ws['B9'].value)
config.array_EH.append(ws['B19'].value)
config.array_EH.append(ws['B29'].value)
config.array_EH.append(ws['B39'].value)

#n HH acquisition
config.array_HH.append(ws['B8'].value)
config.array_HH.append(ws['B18'].value)
config.array_HH.append(ws['B28'].value)
config.array_HH.append(ws['B38'].value)

wb.close()


if config.n_PI != '' and config.n_PI != 0:
# energy house
    gen_EH05.gen_BB_EH05.file_creation_1('./template_excel/template_BB_samsung.xlsx')
    gen_EH05.gen_BR_EH05.file_creation_6('./template_excel/template_BR_samsung.xlsx')
    gen_EH05.gen_PEMS_EH05.file_creation_3('./template_excel/template_PEMS_EH05.xlsx')
    gen_EH05.gen_HVAC_EH05.file_creation_4('./template_excel/template_HVAC_EH05.xlsx')
    
# power house
gen_PH2.gen_INV_PH2.file_creation_0('./template_excel/template_CL_CBESSHD.xlsx')
gen_PH2.gen_AUX_PH2.file_creation_2('./template_excel/template_AUX_CBESSHD.xlsx')
gen_PH2.gen_PEMS_PH2.file_creation_3('./template_excel/template_PEMS_PH2HD.xlsx')
gen_PH2.gen_SKID_PH2.file_creation_5('./template_excel/template_SKID_PH2HD.xlsx')

config.file_aggregation(config.global_list)
config.file_removal(config.global_list)
config.file_reorder('list_alarm.xlsx','new_list_alarm.xlsx')
config.file_header('new_list_alarm.xlsx')
config.file_numbering('new_list_alarm.xlsx')
os.remove('list_alarm.xlsx')
print ('Program ended!')




# config.num_HH_GUI = number_of_HH.get()